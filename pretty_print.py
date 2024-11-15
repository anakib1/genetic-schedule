from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from collections import defaultdict
import csv
from models import *

NUM_SLOTS = 20
WEEKS = 14


def load_groups(filename):
    groups = []
    with open(filename, newline='', encoding='UTF-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            group = Group(
                name=row['name'],
                num_students=int(row['num_students']),
                subjects=row['subjects'].split(';'),
                subgroups=row['subgroups'].split(';')
            )
            groups.append(group)
    return groups


def load_lecturers(filename):
    lecturers = []
    with open(filename, newline='', encoding='UTF-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            subjects_can_teach = {}
            subjects = row['subjects_can_teach'].split(';')
            for subj in subjects:
                name, types = subj.split(':')
                subjects_can_teach[name] = types.split(',')
            lecturer = Lecturer(
                name=row['name'],
                subjects_can_teach=subjects_can_teach
            )
            lecturers.append(lecturer)
    return lecturers


def load_rooms(filename):
    rooms = []
    with open(filename, newline='', encoding='UTF-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            room = Room(
                name=row['name'],
                capacity=int(row['capacity'])
            )
            rooms.append(room)
    return rooms


def load_subjects(filename):
    subjects = {}
    with open(filename, newline='', encoding='UTF-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            subject = Subject(
                name=row['name'],
                total_hours=float(row['total_hours']),
                lecture_hours=float(row['lecture_hours']),
                practical_hours=float(row['practical_hours']),
                needs_subgroup=(row['needs_subgroup'] == 'True')
            )
            subjects[row['name']] = subject
    return subjects


def export_schedule_to_excel(schedule, groups, filename='schedule.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule'

    days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    periods = ['P1', 'P2', 'P3', 'P4']
    time_slots = []
    for day in days:
        for period in periods:
            time_slots.append(f"{day} {period}")

    ws.cell(row=1, column=1, value='Group')
    for col_num, time_slot in enumerate(time_slots, start=2):
        cell = ws.cell(row=1, column=col_num, value=time_slot)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')

    for row_num, group in enumerate(groups, start=2):
        ws.cell(row=row_num, column=1, value=group.name)
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')

        group_schedule = ['' for _ in range(NUM_SLOTS)]

        for slot_index, classes in enumerate(schedule.slots):
            for cls in classes:

                if group.name in cls['groups'] or any(subgroup in cls['groups'] for subgroup in group.subgroups):
                    class_info = f"{cls['subject']} ({cls['type']})\nLecturer: {cls['lecturer']}\nRoom: {cls['room']}"
                    group_schedule[slot_index] = class_info
                    break

        for col_num, class_info in enumerate(group_schedule, start=2):
            if class_info:
                cell = ws.cell(row=row_num, column=col_num, value=class_info)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:

                pass

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min((max_length + 2), 20)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)
    print(f"Schedule exported to {filename}")


def get_main_group_name(subgroup_name, groups):
    for group in groups:
        if subgroup_name == group.name or subgroup_name in group.subgroups:
            return group.name
    return None


def print_subject_hours_report(schedule, groups, subjects):
    group_subject_hours = {group.name: defaultdict(float) for group in groups}

    for slot in schedule.slots:
        for cls in slot:

            class_duration = 1.5
            subject_name = cls['subject']
            class_type = cls['type']
            groups_in_class = cls['groups']

            for group_name in groups_in_class:

                main_group_name = get_main_group_name(group_name, groups)
                if main_group_name:
                    group_subject_hours[main_group_name][subject_name] += class_duration

    print("\nSubject Hours Report:")
    for group in groups:
        print(f"\nGroup {group.name}:")
        for subject_name in group.subjects:
            scheduled_hours = group_subject_hours[group.name].get(subject_name, 0) * WEEKS
            required_hours = subjects[subject_name].total_hours
            if scheduled_hours < required_hours:
                status = 'UNDERREPRESENTED'
            else:
                status = 'OK'
            print(f"  Subject: {subject_name}")
            print(f"    Scheduled Hours: {scheduled_hours}")
            print(f"    Required Hours: {required_hours}")
            if status == 'UNDERREPRESENTED':
                print(f"    Status: {status} *")
            else:
                print(f"    Status: {status}")

    print('\nLoss: ', schedule.fitness)
